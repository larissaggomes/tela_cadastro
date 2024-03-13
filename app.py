import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook


# setando a aparencia padrao do sistema
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    # função que guarda a aparencia e o layout do sistema
    def __init__(self):
        super().__init__()     
        self.layout_config()
        self.appearence()
        self.todo_sistema()

    # função do layout do sistema tamanho e texto
    def layout_config(self):
        self.title("Sistema de Gestão de Clientes")
        self.geometry("700x500")

    # função da aparencia do sistema 
    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000','#fff']).place(x=50, y=430) 
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", 'System'], command=self.change_apm).place(x=50, y=460)

    # função que faz retangulo azul
    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="teal")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Sistema de Gestao de Cliente", font=("Century Gothic bold", 24), text_color="#fff").place(x=190, y=10)
        
        span = ctk.CTkLabel(self, text="Por Favor, preencha todos os campos do formulários!", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=50, y=70)

        ficheiro = pathlib.Path("Cliente.xlsx")

        if ficheiro.exists():
            pass
        else:
            ficheiro=Workbook()
            folha=ficheiro.active
            folha['A1']="Nome Completo"
            folha['B1']="Contato"
            folha['C1']="Idade"
            folha['D1']="Genero"
            folha['E1']="Enderco"
            folha['F1']="Observações"

            ficheiro.save("Clientes.xlsx")

        # função de salvar dados
        def submit():
            
            # pegando os dados dos entrys
            name = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            gender = gender_combobox.get()
            address = address_value.get()
            obs = obs_entry.get(0.0, END)

            if(name=="" or contact=="" or age=="" or address==""):
                messagebox.showerror("Sistema", "ERRO!\nPor favor preencha todos os campos!")
            else:
                ficheiro = openpyxl.load_workbook('Clientes.xlsx')
                folha = ficheiro.active
                folha.cell(column=1, row=folha.max_row+1, value=name)
                folha.cell(column=2, row=folha.max_row, value=contact)
                folha.cell(column=3, row=folha.max_row, value=age)
                folha.cell(column=4, row=folha.max_row, value=gender)
                folha.cell(column=5, row=folha.max_row, value=address)
                folha.cell(column=6, row=folha.max_row, value=obs)

                ficheiro.save(r"Clientes.xlsx")
                messagebox.showinfo("Sistema", "Dados salvos com sucesso")

        # função de limpar dados
        def clear():
            name_value.get("")
            contact_value.get("")
            age_value.get("")
            gender_combobox.get("")
            address_value.get("")
            obs_entry.delete(0.0, END)

        # texts variables
        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()

        # Entrys
        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gohtic bold", 16), fg_color="transparent")
        contact_entry = ctk.CTkEntry(self, width=200, textvariable=contact_value, font=("Century Gohtic", 16), fg_color="transparent")
        age_entry = ctk.CTkEntry(self, width=150, textvariable=age_value, font=("Century Gohtic", 16), fg_color="transparent")
        address_entry = ctk.CTkEntry(self, width=200, textvariable=address_value, font=("Century Gohtic", 16), fg_color="transparent")

        # combobox
        gender_combobox = ctk.CTkComboBox(self, values=["Masculino", "Ferminino"], font=("Century Gothic bold", 14), width=150)
        gender_combobox.set("Masculino")

        # Entrade de observecoes
        obs_entry = ctk.CTkTextbox(self, width=470, height=150, font=("arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")

        # labels
        lb_name = ctk.CTkLabel(self, text="Nome Completo:", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_contact = ctk.CTkLabel(self, text="Contato", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_age = ctk.CTkLabel(self, text="Idade", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_gender = ctk.CTkLabel(self, text="Genero", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_address = ctk.CTkLabel(self, text="Endereço", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_obs = ctk.CTkLabel(self, text="observações", font=("Century Gothic bold", 16), text_color=["#000","#fff"])

        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=300, y=420)
        btn_delete = ctk.CTkButton(self, text="Limpar Dados".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=500, y=420)

        # Posicionando os elementos na janela
        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contact.place(x=450, y=120)
        contact_entry.place(x=450, y=150)

        lb_age.place(x=300, y=190)
        age_entry.place(x=300,y=220)

        lb_gender.place(x=500, y=190)
        gender_combobox.place(x=500, y=220)

        lb_address.place(x=50, y=190)
        address_entry.place(x=50, y=220)

        lb_obs.place(x=50, y=260)
        obs_entry.place(x=150, y=260)

    # função que muda aparencia do sistema
    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

# o que inicia a janela do sistema
if __name__=="__main__":
    app = App()
    app.mainloop()
